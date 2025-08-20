import pandas as pd
import os
import sys
from openpyxl import load_workbook

# Validate and extract match number from command-line argument
if len(sys.argv) < 2:
    raise ValueError("❌ Match number not provided.")
match_number = int(sys.argv[1])

# Path to Excel sheet mounted inside Docker
excel_path = "/app/data/SquadPlayerNames_IndianT20League.xlsx"
if not os.path.exists(excel_path):
    raise FileNotFoundError(f"❌ Excel file not found at {excel_path}")

# Load available sheet names
wb = load_workbook(excel_path, read_only=True)
available_sheets = wb.sheetnames
wb.close()

# Attempt to read available sheets up to the given match number
dfs = []
missing_sheets = []

for i in range(1, match_number + 1):
    sheet_name = f"Match_{i}"
    if sheet_name in available_sheets:
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            df['Match No'] = i  # ✅ Add Match No
            dfs.append(df)
        except Exception as e:
            print(f"⚠️ Error reading sheet {sheet_name}: {e}")
    else:
        missing_sheets.append(sheet_name)

if missing_sheets:
    print(f"⚠️ Skipped missing sheets: {missing_sheets}")

if not dfs:
    raise ValueError("❌ No valid match sheets found to merge.")

merged_df = pd.concat(dfs, ignore_index=True)

# Ensure required columns exist
required_columns = ['Player Name', 'Credits', 'Player Type', 'Team', 'IsPlaying', 'Match No']
if not all(col in merged_df.columns for col in required_columns):
    raise ValueError("❌ Missing required columns in the merged data.")

# Save merged file
merged_df = merged_df.sort_values(by='Player Name')
merged_df = merged_df.drop_duplicates(subset='Player Name', keep='first')
os.makedirs("scripts/data", exist_ok=True)
merged_df.to_csv("scripts/data/match_data.csv", index=False)
print("✅ Merging completed. File saved as 'scripts/data/match_data.csv'.")
